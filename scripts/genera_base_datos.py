# -*- coding: utf-8 -*-
"""Genera la gran base de datos del análisis: una fila por comentario (1,862)
con datos del autor, post de origen y polaridad clasificada con IA.

Requiere que exista data/polaridad/<clave>.json (un array de strings alineado
por índice con data/comments_by_post/<clave>.json).

VALIDACIÓN DE CUADRE: los conteos de polaridad por post deben coincidir
exactamente con el sentimiento agregado de data/analysis/<clave>.json
(las cifras publicadas en el reporte HTML). Si no cuadran, el script aborta.

Salidas (con datos personales — NO subir al repo público, están en .gitignore):
  data/base_datos_comentarios.csv   (UTF-8 con BOM, para Excel)
  data/base_datos_comentarios.xlsx  (con autofiltro y anchos de columna)
"""
import json, os, sys, csv, collections
from datetime import datetime, timedelta

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "data")

meta = json.load(open(os.path.join(DATA, "posts_metadata.json"), encoding="utf-8"))
raw = json.load(open(os.path.join(DATA, "comments_raw.json"), encoding="utf-8"))
consolidado = json.load(open(os.path.join(DATA, "consolidado.json"), encoding="utf-8"))

# mapa inputUrl -> clave (misma lógica que prepara_datos.py para garantizar
# el mismo orden por post y la misma asignación)
url_a_clave = {}
for p in meta["posts"]:
    url_a_clave[p["url"].rstrip("/")] = p["clave"]
    if "url_canonica" in p:
        url_a_clave[p["url_canonica"].rstrip("/")] = p["clave"]

info_post = {p["clave"]: p for p in consolidado["posts"]}
orden_post = {p["clave"]: i for i, p in enumerate(consolidado["posts"])}

# polaridades por post (alineadas por índice con comments_by_post)
polaridad = {}
for clave in info_post:
    ruta = os.path.join(DATA, "polaridad", clave + ".json")
    if not os.path.exists(ruta):
        sys.exit(f"FALTA {ruta} — corre primero la clasificación por post.")
    polaridad[clave] = json.load(open(ruta, encoding="utf-8"))

def fecha_local_bcs(iso):
    """UTC -> hora local BCS (UTC-7), formato legible."""
    if not iso:
        return ""
    dt = datetime.strptime(iso[:19], "%Y-%m-%dT%H:%M:%S") - timedelta(hours=7)
    return dt.strftime("%Y-%m-%d %H:%M")

filas = []
idx_por_post = collections.Counter()
for c in raw:
    clave = url_a_clave.get((c.get("inputUrl") or "").rstrip("/"))
    if not clave:
        continue
    i = idx_por_post[clave]
    idx_por_post[clave] += 1
    etiquetas = polaridad[clave]
    if i >= len(etiquetas):
        sys.exit(f"DESALINEACIÓN en {clave}: comentario {i} sin etiqueta "
                 f"(el archivo de polaridad tiene {len(etiquetas)}).")
    pid = c.get("profileId") or ""
    filas.append({
        "post_clave": clave,
        "post_titulo": info_post[clave]["titulo"],
        "post_url": info_post[clave]["url"],
        "fecha_utc": c.get("date") or "",
        "fecha_local_bcs": fecha_local_bcs(c.get("date")),
        "autor": c.get("profileName") or "",
        "autor_id": pid,
        "autor_perfil_url": f"https://www.facebook.com/{pid}" if pid else "",
        "comentario": (c.get("text") or "").strip(),
        "likes": c.get("likesCount") or 0,
        "es_respuesta": "Sí" if (c.get("threadingDepth") or 0) > 0 else "No",
        "respuestas": c.get("commentsCount") or 0,
        "polaridad": etiquetas[i],
        "comentario_url": c.get("commentUrl") or "",
        "comentario_id": c.get("commentId") or "",
    })

# ---------- validación de cuadre con el reporte HTML ----------
errores = []
conteo = collections.defaultdict(collections.Counter)
for f in filas:
    conteo[f["post_clave"]][f["polaridad"]] += 1
# "sin texto": comentarios de solo sticker/adjunto que el análisis publicado
# excluyó de la base de sentimiento (caso reel_claro_que_se_puede: 306 de 319)
print(f"{'post':<26}{'pos':>6}{'neg':>6}{'neu':>6}{'s/txt':>7}{'total':>7}  cuadre")
for clave, p in info_post.items():
    esperado = p["sentimiento"]  # cifras publicadas en el reporte
    obtenido = conteo[clave]
    ok = all(obtenido.get(k, 0) == esperado[k] for k in ("positivo", "negativo", "neutro"))
    print(f"{clave:<26}{obtenido.get('positivo',0):>6}{obtenido.get('negativo',0):>6}"
          f"{obtenido.get('neutro',0):>6}{obtenido.get('sin texto',0):>7}"
          f"{sum(obtenido.values()):>7}  {'OK' if ok else 'X NO CUADRA'}")
    if not ok:
        errores.append(f"{clave}: esperado {esperado}, obtenido {dict(obtenido)}")
if len(filas) != len(raw):
    print(f"AVISO: {len(raw) - len(filas)} comentarios del crudo sin match de post.")
if errores:
    sys.exit("NO CUADRA con el reporte HTML:\n" + "\n".join(errores))
print(f"\nTotal filas: {len(filas)} — todo cuadra con el reporte HTML.")

# orden: mismo orden de posts que el reporte, luego cronológico
filas.sort(key=lambda f: (orden_post[f["post_clave"]], f["fecha_utc"]))

COLS = ["post_clave", "post_titulo", "post_url", "fecha_utc", "fecha_local_bcs",
        "autor", "autor_id", "autor_perfil_url", "comentario", "likes",
        "es_respuesta", "respuestas", "polaridad", "comentario_url", "comentario_id"]

# ---------- CSV (UTF-8 con BOM para que Excel lo abra bien) ----------
ruta_csv = os.path.join(DATA, "base_datos_comentarios.csv")
with open(ruta_csv, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=COLS)
    w.writeheader()
    w.writerows(filas)
print("Escrito:", ruta_csv)

# ---------- XLSX ----------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl no disponible; solo se generó el CSV.")

wb = Workbook()
ws = wb.active
ws.title = "Comentarios"
ws.append(COLS)
morado = PatternFill("solid", fgColor="684D9B")
for celda in ws[1]:
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = morado
for fila in filas:
    ws.append([fila[c] for c in COLS])
anchos = {"post_clave": 22, "post_titulo": 34, "post_url": 18, "fecha_utc": 20,
          "fecha_local_bcs": 17, "autor": 24, "autor_id": 18, "autor_perfil_url": 18,
          "comentario": 70, "likes": 7, "es_respuesta": 12, "respuestas": 11,
          "polaridad": 11, "comentario_url": 18, "comentario_id": 18}
for i, c in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = anchos[c]
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"
ruta_xlsx = os.path.join(DATA, "base_datos_comentarios.xlsx")
wb.save(ruta_xlsx)
print("Escrito:", ruta_xlsx)
